#!/usr/bin/env python3
"""Generate QA Test Cases Excel for La Tabla project."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = openpyxl.Workbook()

# Styles
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid")
section_font = Font(bold=True, color="FFFFFF", size=11)
section_fill = PatternFill(start_color="B8860B", end_color="B8860B", fill_type="solid")
pass_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
fail_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
blocked_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
wrap = Alignment(wrap_text=True, vertical="top")

HEADERS = ["ID", "Módulo", "Caso de Prueba", "Precondiciones", "Pasos", "Resultado Esperado", "Prioridad", "Estado", "Tester", "Fecha", "Observaciones"]
COL_WIDTHS = [8, 16, 30, 25, 40, 30, 12, 12, 12, 12, 30]

def setup_sheet(ws, title):
    ws.title = title
    ws.sheet_properties.tabColor = "B8860B"
    for i, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:K1"

def add_cases(ws, cases, start_row=2):
    row = start_row
    for case in cases:
        for i, val in enumerate(case, 1):
            cell = ws.cell(row=row, column=i, value=val)
            cell.alignment = wrap
            cell.border = thin_border
        row += 1
    return row

# ============================================================
# SHEET 1: Navegación y UI
# ============================================================
ws1 = wb.active
setup_sheet(ws1, "Navegación y UI")
add_cases(ws1, [
    ["NAV-001", "Landing", "Carga de página principal", "Servidor activo", "1. Abrir https://la-tabla.vercel.app/\n2. Esperar carga completa", "Landing carga con hero, productos destacados, FAQ, zonas de entrega, footer", "Alta", "", "", "", ""],
    ["NAV-002", "Landing", "Productos destacados muestran datos reales", "Productos activos en DB", "1. Abrir landing\n2. Verificar sección de productos destacados", "Muestra top 3 tablas y top 3 especialidades con nombre, precio, imagen", "Alta", "", "", "", ""],
    ["NAV-003", "Navbar", "Navegación entre secciones", "Página cargada", "1. Click en cada link del navbar: Tablas, Especialidades, Servicios, Talleres, Nosotros, Contacto, Mi Pedido", "Cada link lleva a su página correcta", "Alta", "", "", "", ""],
    ["NAV-004", "Navbar", "Search bar funcional", "Productos en DB", "1. Click en icono de búsqueda\n2. Escribir 'tabla'\n3. Esperar resultados", "Muestra resultados filtrados en tiempo real (debounce 300ms)", "Alta", "", "", "", ""],
    ["NAV-005", "Navbar", "Search sin resultados", "Página cargada", "1. Buscar 'xyzabc123'", "Muestra mensaje de sin resultados", "Media", "", "", "", ""],
    ["NAV-006", "Navbar", "Icono admin visible solo para admin", "Logueado como admin", "1. Verificar navbar logueado como admin\n2. Verificar navbar sin login", "Icono de dashboard visible solo para admin", "Alta", "", "", "", ""],
    ["NAV-007", "Footer", "Links del footer funcionan", "Página cargada", "1. Click en cada link del footer: Sobre nosotros, Contacto, Seguí tu pedido, etc.", "Todos los links llevan a sus páginas", "Media", "", "", "", ""],
    ["NAV-008", "404", "Ruta inexistente redirige al landing", "Servidor activo", "1. Navegar a /ruta-que-no-existe", "Redirige automáticamente al landing page", "Media", "", "", "", ""],
    ["NAV-009", "Responsive", "Mobile - navbar hamburger", "Viewport < 768px", "1. Abrir en móvil o reducir ventana\n2. Verificar menú hamburger\n3. Abrir menú", "Menú hamburger funcional, links accesibles", "Alta", "", "", "", ""],
    ["NAV-010", "Responsive", "Mobile - landing page", "Viewport < 768px", "1. Abrir landing en móvil\n2. Scroll completo", "Todo el contenido se ve correctamente, sin overflow horizontal", "Alta", "", "", "", ""],
    ["NAV-011", "SEO", "Meta tags presentes", "Página cargada", "1. Inspeccionar <head> del HTML\n2. Verificar title, description, og:image", "Title, description y OG tags presentes y correctos", "Media", "", "", "", ""],
    ["NAV-012", "SEO", "Sitemap accesible", "Servidor activo", "1. Abrir /sitemap.xml", "Sitemap XML válido con todas las rutas", "Baja", "", "", "", ""],
    ["NAV-013", "SEO", "Robots.txt accesible", "Servidor activo", "1. Abrir /robots.txt", "Robots.txt válido permitiendo crawling", "Baja", "", "", "", ""],
    ["NAV-014", "Loading", "Skeleton loading se muestra", "Conexión lenta o throttle", "1. Throttle network a Slow 3G\n2. Navegar a /tablas", "Skeletons animados mientras carga", "Media", "", "", "", ""],
    ["NAV-015", "Contacto", "Página de contacto muestra info", "Servidor activo", "1. Abrir /contacto", "Muestra Instagram, email, horario, zonas de cobertura. NO WhatsApp", "Alta", "", "", "", ""],
    ["NAV-016", "Sobre Nosotros", "Página about carga", "Servidor activo", "1. Abrir /sobre-nosotros", "Historia del chef, valores, servicios, CTA", "Media", "", "", "", ""],
])

# ============================================================
# SHEET 2: Catálogo de Productos
# ============================================================
ws2 = wb.create_sheet()
setup_sheet(ws2, "Catálogo")
add_cases(ws2, [
    ["CAT-001", "Tablas", "Listado de tablas", "Tablas activas en DB", "1. Navegar a /tablas", "Grid con todas las tablas activas, imagen, nombre, precio", "Alta", "", "", "", ""],
    ["CAT-002", "Tablas", "Detalle de tabla fija", "Tabla no configurable en DB", "1. Click en una tabla fija", "Muestra galería, nombre, precio, descripción, ingredientes fijos, botón agregar al carrito", "Alta", "", "", "", ""],
    ["CAT-003", "Tablas", "Detalle de tabla configurable", "Tabla configurable con reglas", "1. Click en tabla configurable\n2. Verificar builder", "Muestra builder con categorías, cantidad requerida por categoría, ingredientes seleccionables", "Alta", "", "", "", ""],
    ["CAT-004", "Tablas", "Builder - selección de ingredientes", "En detalle de tabla configurable", "1. Seleccionar ingredientes según reglas\n2. Verificar contador", "Contador muestra X/Y seleccionados, no permite exceder cantidad", "Alta", "", "", "", ""],
    ["CAT-005", "Tablas", "Builder - agregar al carrito con ingredientes", "Ingredientes seleccionados correctamente", "1. Completar selección\n2. Click agregar al carrito", "Producto se agrega con ingredientes seleccionados, toast de confirmación", "Alta", "", "", "", ""],
    ["CAT-006", "Especialidades", "Listado de especialidades", "Especialidades activas en DB", "1. Navegar a /especialidades", "Grid con especialidades activas", "Alta", "", "", "", ""],
    ["CAT-007", "Especialidades", "Detalle de especialidad", "Especialidad en DB", "1. Click en una especialidad", "Galería/imagen, nombre, precio, descripción, botón agregar", "Alta", "", "", "", ""],
    ["CAT-008", "Servicios", "Listado de servicios", "Servicios activos en DB", "1. Navegar a /servicios", "Grid con servicios activos", "Alta", "", "", "", ""],
    ["CAT-009", "Servicios", "Detalle de servicio", "Servicio en DB", "1. Click en un servicio", "Info completa con personas min/max si aplica", "Alta", "", "", "", ""],
    ["CAT-010", "Talleres", "Listado de talleres", "Talleres activos en DB", "1. Navegar a /talleres", "Grid con talleres activos", "Alta", "", "", "", ""],
    ["CAT-011", "Talleres", "Detalle de taller", "Taller en DB", "1. Click en un taller", "Info completa del taller", "Alta", "", "", "", ""],
    ["CAT-012", "Vacío", "Categoría sin productos muestra empty state", "Categoría sin productos activos", "1. Desactivar todos los productos de una categoría\n2. Navegar a su listado", "Muestra 'Próximamente' con icono", "Media", "", "", "", ""],
    ["CAT-013", "Galería", "Galería de producto con múltiples imágenes", "Producto con 3+ imágenes", "1. Abrir detalle del producto\n2. Navegar entre imágenes", "Galería funcional, thumbnails clickeables, imagen principal cambia", "Media", "", "", "", ""],
    ["CAT-014", "JSON-LD", "Schema.org en detalle producto", "Producto con precio", "1. Inspeccionar HTML del detalle\n2. Buscar script type=application/ld+json", "JSON-LD válido con name, price, currency USD", "Baja", "", "", "", ""],
])

# ============================================================
# SHEET 3: Carrito y Checkout
# ============================================================
ws3 = wb.create_sheet()
setup_sheet(ws3, "Carrito y Checkout")
add_cases(ws3, [
    ["CHK-001", "Carrito", "Agregar producto al carrito", "Producto disponible", "1. Ir a detalle de producto\n2. Click 'Agregar al carrito'", "Producto aparece en carrito, toast de confirmación, badge del carrito se actualiza", "Alta", "", "", "", ""],
    ["CHK-002", "Carrito", "Agregar mismo producto incrementa cantidad", "Producto ya en carrito", "1. Agregar mismo producto 2 veces", "Cantidad incrementa a 2, no duplica item", "Alta", "", "", "", ""],
    ["CHK-003", "Carrito", "Cambiar cantidad en carrito", "Items en carrito", "1. Ir a /carrito\n2. Cambiar cantidad con +/-", "Cantidad y subtotal se actualizan", "Alta", "", "", "", ""],
    ["CHK-004", "Carrito", "Eliminar item del carrito", "Items en carrito", "1. Click en eliminar item", "Item desaparece, total se recalcula", "Alta", "", "", "", ""],
    ["CHK-005", "Carrito", "Carrito vacío", "Sin items", "1. Ir a /carrito sin items", "Mensaje de carrito vacío con link a catálogo", "Media", "", "", "", ""],
    ["CHK-006", "Carrito", "Persistencia en localStorage", "Items en carrito", "1. Agregar items\n2. Cerrar pestaña\n3. Reabrir", "Items persisten después de cerrar y reabrir", "Alta", "", "", "", ""],
    ["CHK-007", "Checkout", "Formulario completo - delivery", "Items en carrito", "1. Ir a checkout\n2. Llenar: nombre, email, teléfono\n3. Seleccionar delivery\n4. Llenar dirección\n5. Seleccionar fecha y horario\n6. Seleccionar método de pago\n7. Enviar", "Pedido se crea exitosamente, redirige a confirmación", "Alta", "", "", "", ""],
    ["CHK-008", "Checkout", "Formulario completo - pickup", "Items en carrito", "1. Seleccionar pickup\n2. Completar datos\n3. Enviar", "Pedido se crea sin dirección requerida", "Alta", "", "", "", ""],
    ["CHK-009", "Checkout", "Validación de campos requeridos", "Checkout abierto", "1. Intentar enviar formulario vacío", "Muestra errores de validación en campos requeridos", "Alta", "", "", "", ""],
    ["CHK-010", "Checkout", "Email inválido rechazado", "Checkout abierto", "1. Ingresar email inválido\n2. Enviar", "Muestra error de formato de email", "Alta", "", "", "", ""],
    ["CHK-011", "Checkout", "Fecha bloqueada no seleccionable", "Fecha bloqueada en DB", "1. Abrir date picker\n2. Intentar seleccionar fecha bloqueada", "Fecha aparece deshabilitada, no se puede seleccionar", "Alta", "", "", "", ""],
    ["CHK-012", "Checkout", "Fecha mínima 2 días adelante", "Checkout abierto", "1. Abrir date picker\n2. Intentar seleccionar hoy o mañana", "Fechas anteriores a 2 días aparecen deshabilitadas", "Alta", "", "", "", ""],
    ["CHK-013", "Checkout", "Delivery slots cargan de DB", "Slots en DB", "1. Seleccionar fecha válida", "Muestra horarios disponibles del DB", "Alta", "", "", "", ""],
    ["CHK-014", "Confirmación", "Página de confirmación", "Pedido recién creado", "1. Completar checkout exitoso", "Muestra número de pedido, resumen, link a tracking", "Alta", "", "", "", ""],
    ["CHK-015", "Email", "Email de confirmación al cliente", "Pedido creado con email válido", "1. Crear pedido\n2. Verificar inbox", "Cliente recibe email con resumen del pedido", "Alta", "", "", "", ""],
    ["CHK-016", "Email", "Email de notificación al admin", "Pedido creado", "1. Crear pedido\n2. Verificar inbox admin", "Admin recibe notificación con detalles del pedido", "Alta", "", "", "", ""],
    ["CHK-017", "Rate Limit", "Rate limiting en orders", "Servidor activo", "1. Enviar 6 pedidos en menos de 1 minuto desde la misma IP", "Después del 5to, responde 429 Too Many Requests", "Media", "", "", "", ""],
])

# ============================================================
# SHEET 4: Tracking de Pedidos
# ============================================================
ws4 = wb.create_sheet()
setup_sheet(ws4, "Tracking")
add_cases(ws4, [
    ["TRK-001", "Tracking", "Buscar pedido por número", "Pedido existente", "1. Ir a /pedido\n2. Ingresar número de pedido válido", "Muestra estado actual, barra de progreso, detalles del pedido", "Alta", "", "", "", ""],
    ["TRK-002", "Tracking", "Pedido no encontrado", "Número inexistente", "1. Ingresar número falso 'LT-000000-XXXX'", "Muestra 'Pedido no encontrado'", "Alta", "", "", "", ""],
    ["TRK-003", "Tracking", "Barra de progreso refleja estado", "Pedidos en distintos estados", "1. Verificar tracking de pedido pending\n2. Verificar tracking de pedido confirmed\n3. etc.", "Barra de progreso resalta el estado actual correctamente", "Alta", "", "", "", ""],
    ["TRK-004", "Tracking", "PII limitado", "Pedido existente", "1. Ver tracking\n2. Verificar datos mostrados", "Solo muestra primer nombre, NO email, teléfono ni dirección completa", "Alta", "", "", "", ""],
    ["TRK-005", "Tracking", "Auto-fill desde confirmación", "Pedido recién creado", "1. Completar checkout\n2. Click en 'Seguir pedido' desde confirmación", "Número de pedido se auto-llena y muestra el tracking", "Media", "", "", "", ""],
    ["TRK-006", "Tracking", "Rate limiting", "Servidor activo", "1. Hacer 11 búsquedas en menos de 1 minuto", "Después del 10mo, responde 429", "Media", "", "", "", ""],
])

# ============================================================
# SHEET 5: Admin Panel
# ============================================================
ws5 = wb.create_sheet()
setup_sheet(ws5, "Admin Panel")
add_cases(ws5, [
    ["ADM-001", "Auth", "Acceso admin requiere login", "No logueado", "1. Navegar a /admin", "Redirige a /sign-in", "Alta", "", "", "", ""],
    ["ADM-002", "Auth", "Usuario no-admin no puede acceder", "Logueado con email no-admin", "1. Navegar a /admin", "Redirige o muestra error de permisos", "Alta", "", "", "", ""],
    ["ADM-003", "Auth", "Admin accede correctamente", "Logueado como dvargas.dev@gmail.com", "1. Navegar a /admin", "Dashboard carga con estadísticas", "Alta", "", "", "", ""],
    ["ADM-004", "Dashboard", "Stats muestran datos reales", "Pedidos y productos en DB", "1. Verificar tarjetas del dashboard", "Muestra: pedidos hoy, ingresos, pedidos pendientes, productos activos", "Alta", "", "", "", ""],
    ["ADM-005", "Productos", "Listar productos", "Productos en DB", "1. Ir a /admin/productos", "Lista todos los productos con nombre, tipo, precio, estado", "Alta", "", "", "", ""],
    ["ADM-006", "Productos", "Crear producto", "En /admin/productos", "1. Click 'Nuevo Producto'\n2. Llenar formulario\n3. Guardar", "Producto se crea y aparece en la lista", "Alta", "", "", "", ""],
    ["ADM-007", "Productos", "Editar producto", "Producto existente", "1. Click editar en un producto\n2. Cambiar nombre y precio\n3. Guardar", "Cambios se reflejan en lista y en el sitio público", "Alta", "", "", "", ""],
    ["ADM-008", "Productos", "Desactivar producto", "Producto activo", "1. Editar producto\n2. Desactivar\n3. Guardar", "Producto no aparece en el sitio público", "Alta", "", "", "", ""],
    ["ADM-009", "Productos", "Slug se genera automáticamente", "Creando producto", "1. Escribir nombre 'Tabla Española'\n2. Verificar slug", "Slug auto-generado: 'tabla-espanola'", "Media", "", "", "", ""],
    ["ADM-010", "Productos", "Upload de imagen", "Creando/editando producto", "1. Click en upload imagen\n2. Seleccionar archivo\n3. Guardar", "Imagen se sube a Vercel Blob y se muestra", "Alta", "", "", "", ""],
    ["ADM-011", "Ingredientes", "Listar ingredientes", "Ingredientes en DB", "1. Ir a /admin/ingredientes", "Lista ingredientes con nombre, categoría, disponibilidad", "Alta", "", "", "", ""],
    ["ADM-012", "Ingredientes", "Crear ingrediente", "Categorías existentes", "1. Click nuevo\n2. Nombre, categoría, costo\n3. Guardar", "Ingrediente creado y visible en lista", "Alta", "", "", "", ""],
    ["ADM-013", "Ingredientes", "Marcar no disponible", "Ingrediente disponible", "1. Editar ingrediente\n2. Desmarcar disponible\n3. Guardar", "Ingrediente no aparece en el builder de tablas", "Alta", "", "", "", ""],
    ["ADM-014", "Categorías", "CRUD categorías", "En /admin/categorias", "1. Crear categoría 'Frutas'\n2. Editar nombre\n3. Intentar eliminar", "Crear/editar funciona. Eliminar falla si tiene ingredientes asociados", "Alta", "", "", "", ""],
    ["ADM-015", "Pedidos", "Ver lista de pedidos", "Pedidos en DB", "1. Ir a /admin/pedidos", "Lista de pedidos ordenados por fecha, con estado y total", "Alta", "", "", "", ""],
    ["ADM-016", "Pedidos", "Actualizar estado de pedido", "Pedido pendiente", "1. Click en pedido\n2. Cambiar estado a 'Confirmado'\n3. Guardar", "Estado se actualiza correctamente", "Alta", "", "", "", ""],
    ["ADM-017", "Pedidos", "Verificar pago", "Pedido con pago pendiente", "1. Cambiar paymentStatus a 'Verificado'\n2. Guardar", "Estado de pago se actualiza", "Alta", "", "", "", ""],
    ["ADM-018", "Calendario", "Ver pedidos por fecha", "Pedidos con fechas de entrega", "1. Ir a /admin/calendario", "Muestra pedidos agrupados por fecha de entrega", "Media", "", "", "", ""],
    ["ADM-019", "Calendario", "Bloquear fecha", "En /admin/calendario", "1. Agregar fecha bloqueada\n2. Verificar en checkout", "Fecha aparece bloqueada en el calendario y en checkout", "Alta", "", "", "", ""],
    ["ADM-020", "Galería", "Subir imagen a galería", "En /admin/galeria", "1. Click upload\n2. Seleccionar imagen\n3. Asignar a producto", "Imagen sube a Vercel Blob y aparece en galería", "Alta", "", "", "", ""],
    ["ADM-021", "Galería", "Eliminar imagen", "Imagen en galería", "1. Hover sobre imagen\n2. Click eliminar\n3. Confirmar", "Imagen se elimina de galería y Blob", "Media", "", "", "", ""],
    ["ADM-022", "APIs", "API protegida sin auth", "No logueado", "1. POST /api/products con body válido sin auth", "Responde 401 Unauthorized", "Alta", "", "", "", ""],
    ["ADM-023", "APIs", "Validación Zod en APIs", "Logueado como admin", "1. POST /api/products con body inválido (sin name)", "Responde 400 con error descriptivo", "Alta", "", "", "", ""],
])

# ============================================================
# SHEET 6: Resumen
# ============================================================
ws6 = wb.create_sheet()
ws6.title = "Resumen"
ws6.sheet_properties.tabColor = "1A1A1A"

# Header
ws6.merge_cells("A1:F1")
cell = ws6.cell(row=1, column=1, value="La Tabla — QA Test Summary")
cell.font = Font(bold=True, size=16, color="B8860B")
cell.alignment = Alignment(horizontal="center")

summary_headers = ["Módulo", "Total Casos", "Pasados", "Fallidos", "Bloqueados", "% Completado"]
for i, h in enumerate(summary_headers, 1):
    cell = ws6.cell(row=3, column=i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border

modules = [
    ["Navegación y UI", 16, 0, 0, 0, "0%"],
    ["Catálogo", 14, 0, 0, 0, "0%"],
    ["Carrito y Checkout", 17, 0, 0, 0, "0%"],
    ["Tracking", 6, 0, 0, 0, "0%"],
    ["Admin Panel", 23, 0, 0, 0, "0%"],
    ["TOTAL", 76, 0, 0, 0, "0%"],
]

for r, mod in enumerate(modules, 4):
    for c, val in enumerate(mod, 1):
        cell = ws6.cell(row=r, column=c, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
    if mod[0] == "TOTAL":
        for c in range(1, 7):
            ws6.cell(row=r, column=c).font = Font(bold=True)
            ws6.cell(row=r, column=c).fill = PatternFill(start_color="F5F5DC", end_color="F5F5DC", fill_type="solid")

for i, w in enumerate([20, 14, 14, 14, 14, 16], 1):
    ws6.column_dimensions[get_column_letter(i)].width = w

# Instructions
ws6.cell(row=11, column=1, value="Instrucciones:").font = Font(bold=True, size=12)
instructions = [
    "1. Cada pestaña contiene los casos de prueba de un módulo",
    "2. Estado: dejar vacío = pendiente, o escribir PASS / FAIL / BLOCKED / SKIP",
    "3. Registrar nombre del tester y fecha de ejecución",
    "4. Usar columna Observaciones para bugs, screenshots, notas",
    "5. Actualizar la pestaña Resumen manualmente con los conteos",
    "6. Prioridad: Alta = debe funcionar, Media = importante, Baja = nice-to-have",
]
for i, inst in enumerate(instructions, 12):
    ws6.cell(row=i, column=1, value=inst)
    ws6.merge_cells(f"A{i}:F{i}")

# Move Resumen to first position
wb.move_sheet(ws6, offset=-5)

# Save
output = "/home/ubuntu/.openclaw/workspace/la-tabla/docs/QA-La-Tabla.xlsx"
wb.save(output)
print(f"✅ Generated: {output}")
print(f"   Sheets: {len(wb.sheetnames)}")
print(f"   Total test cases: 76")
